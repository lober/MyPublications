#!/usr/bin/env python3
"""
Build a spreadsheet reconciling publications found in NCBI MyBibliography vs OpenAlex.

Reads:
    data/merged_publications.json  (from merge_sources.py)

Writes:
    output/publications_reconciliation.xlsx

Usage:
    python3 scripts/build_reconciliation_spreadsheet.py
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
MERGED_PATH = BASE_DIR / "data" / "merged_publications.json"
OUT_PATH = BASE_DIR / "output" / "publications_reconciliation.xlsx"

FONT_NAME = "Arial"


def main():
    merged = json.loads(MERGED_PATH.read_text())

    wb = Workbook()
    ws = wb.active
    ws.title = "Publications"

    headers = ["Citation", "In NCBI", "In OpenAlex", "In Summary"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name=FONT_NAME, bold=True)

    for row, pub in enumerate(merged, start=2):
        ws.cell(row=row, column=1, value=pub["citation"]).font = Font(name=FONT_NAME)
        ws.cell(row=row, column=2, value=pub["in_ncbi"]).font = Font(name=FONT_NAME)
        ws.cell(row=row, column=3, value=pub["in_openalex"]).font = Font(name=FONT_NAME)
        ws.cell(row=row, column=4, value=pub["in_summary"]).font = Font(name=FONT_NAME)

    ws.column_dimensions[get_column_letter(1)].width = 110
    ws.column_dimensions[get_column_letter(2)].width = 10
    ws.column_dimensions[get_column_letter(3)].width = 12
    ws.column_dimensions[get_column_letter(4)].width = 12
    ws.freeze_panes = "A2"

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Saved {len(merged)} rows to {OUT_PATH}")


if __name__ == "__main__":
    main()
